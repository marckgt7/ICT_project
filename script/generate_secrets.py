import pandas as pd
import re
import os
import random
from openpyxl import Workbook, load_workbook

perfect = 0
imperfect = 0

regex_no_separators = [r'(?i)(?:float)(?:.|[\n\r]){0,40}\b([a-f0-9]{16}[A-Za-z0-9+/]{42,43}\=)',
                       r'\b([a-z]{2}[0-9]{9})\b',
                       r'(?i)(?:front)(?:.|[\n\r]){0,40}\b([0-9a-zA-Z]{36}\.[0-9a-zA-Z.-_]{188,244})\b',
                       r'(?i)(?:image4)(?:.|[\n\r]){0,40}\b([0-9a-zA-Z]{22}[0-9a-zA-Z=]{2})',
                       r'(?i)(?:instabot)(?:.|[\n\r]){0,40}\b([0-9a-zA-Z=+\/]{43}[0-9a-zA-Z+\/=]{1})',
                       r'(?i)(?:instabot)(?:.|[\n\r]){0,40}\b([0-9a-zA-Z=+\/]{43}[0-9a-zA-Z+\/=]{1})',
                       r'(?i)(?:instabot)(?:.|[\n\r]){0,40}\b([0-9a-zA-Z=+]{43}[0-9a-zA-Z+=]{1})']

def aggiungi_regex_in_excel(file_excel, regex):
    try:
        # Controlla se il file Excel esiste
        try:
            workbook = load_workbook(file_excel)
            sheet = workbook.active
        except FileNotFoundError:
            # Crea un nuovo workbook se il file non esiste
            workbook = Workbook()
            sheet = workbook.active
            # Intestazione (opzionale)
            sheet.append(["Regex"])
        
        # Controlla se la regex è già presente
        existing_regexes = [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]
        if regex not in existing_regexes:
            # Aggiunge la regex
            sheet.append([regex])
            workbook.save(file_excel)
            print(f"Regex aggiunta: {regex}")
    except Exception as e:
        print(f"Errore: {e}")


def sanitize_filename(name):
    return re.sub(r'[\\/:*?"<>|]', '_', name)

def expand_charset(charset):
    """
    Expands shorthand character ranges in the input string into full character sequences.
    
    Args:
        charset (str): A string representing a character set with optional ranges (e.g., '0-9a-zA-Z').
    
    Returns:
        str: The expanded character set with all ranges replaced by the full character sequences.
    """
    expanded_charset = ""
    i = 0
    
    while i < len(charset):
        # Check for a valid range pattern (e.g., 'a-z')
        if i + 2 < len(charset) and charset[i + 1] == '-' and charset[i].isalnum() and charset[i + 2].isalnum():
            start, end = charset[i], charset[i + 2]
            
            # Ensure the start and end are in the correct order and are of the same type (both digits or both letters)
            if ord(start) <= ord(end) and (
                (start.isdigit() and end.isdigit()) or 
                (start.islower() and end.islower()) or 
                (start.isupper() and end.isupper())
            ):
                expanded_charset += ''.join(chr(c) for c in range(ord(start), ord(end) + 1))
                i += 3  # Skip over the range (e.g., 'a-z')
            else:
                expanded_charset += charset[i]  # Add the character as-is if the range is invalid
                i += 1
        else:
            expanded_charset += charset[i]  # Add the character as-is
            i += 1
    
    return expanded_charset



def generate_random_string(charset, length):
    global perfect, imperfect
    # Espandi il charset se contiene intervalli
    expanded_charset = expand_charset(charset)
    
    # Se length è un intervallo, scegli una lunghezza casuale nell'intervallo
    if ',' in length:
        min_len, max_len = length.split(',')
        min_len = int(min_len) if min_len else None
        max_len = int(max_len) if max_len else None
        
        if min_len is not None and max_len is not None:
            actual_length = random.randint(min_len, max_len)
        elif min_len is not None:
            actual_length = min_len  # Usa min_len come lunghezza fissa
        elif max_len is not None:
            actual_length = max_len  # Usa max_len come lunghezza fissa
        else:
            imperfect += 1

    elif length == 'None' or length == '\\':
        # Lunghezza casuale se è 'None'
        actual_length = random.randint(1, 100)  # Puoi cambiare l'intervallo
        imperfect += 1
    else:
        # Lunghezza fissa
        actual_length = int(length)
    
    #print(actual_length)
    return ''.join(random.choices(expanded_charset, k=actual_length))

# Funzione per analizzare una singola regex e ottenere informazioni dettagliate
def generate_secret(regex):
    global perfect, imperfect

    if regex == "EMAIL A PIACERE":
       return "testuser.1005@example.com"

    # Trova il pre fisso
    match = re.search(r"\(\?:([a-zA-Z0-9|_.-]+)\)", regex) 
    if match:
        prefisso = match.group(1).split('|')
    else:
        match = re.search(r"\\b\(?([a-zA-Z0-9_.\\-]+)\\?", regex) or \
                re.search(r"\(\?:([a-zA-Z0-9]+)\)", regex) or \
                re.search(r'https://[^[]+', re.sub(r'\\', '', regex))
        if match:
            if match.lastindex is not None and match.lastindex >= 1: 
                prefisso = match.group(1) 
                prefisso = prefisso.replace("\\", "")
            else: 
                prefisso = match.group(0)  # HTTP 
        else:
            prefisso = ""
            aggiungi_regex_in_excel("regex.xlsx", regex)
            imperfect += 1
    
    if(prefisso == "https"): prefisso += '://'

    # Matching per suffissi tipo ".aha.io", ".jfrog.io" non chiusi tra [] o {}
    suffix_match = re.findall(r'(\.[a-zA-Z0-9]+|@[a-zA-Z0-9]+|tt\b|[ \r\n]{1}|-X)', regex)
    if not suffix_match and "eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiJhcGkubGl2ZXN0b3JtLmNvIiwianRpIjoi" not in prefisso:
        suffix_match = re.findall(r'(?:\b|\W)([a-zA-Z0-9]+)(?=\)\b)', regex)

    # Pattern per trovare i separatori (es. \. o \-)
    separator_pattern = r'\\(-4|[.=_|@\-])'  # Cattura -4 come entità e tutti gli altri separatori

    # Estrarre i separatori
    all_separators = []
    matches = re.findall(separator_pattern, regex)
    all_separators.extend(matches)

    if("eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiJhcGkubGl2ZXN0b3JtLmNvIiwianRpIjoi" in prefisso):
        suffix_match.remove(".eyJhdWQiOiJhcGkubGl2ZXN0b3JtLmNvIiwianRpIjoi")

    # Se suffix_match ha qualcosa, rimuove gli ultimi n elementi da all_separators
    if suffix_match and ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9" not in prefisso and "gocanvas" not in prefisso and "intercom" not in prefisso and "eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiJhcGkubGl2ZXN0b3JtLmNvIiwianRpIjoi" not in prefisso):
        n = len(suffix_match)
        all_separators = all_separators[:-n] if n <= len(all_separators) else []
        
    if '=' in all_separators:
        suffix_match.insert(0, '=')
        all_separators.remove('=')

    # Trova tutte le coppie tra parentesi quadre e graffe
    matches = re.findall(r"\[([^\]]+)\]|\{([^\}]+)\}", regex)

    if(".eyJ" in suffix_match or ".ey" in suffix_match):
        all_separators.insert(0, suffix_match[0])  # Usa l'elemento in testa
        suffix_match.pop(0)  # Rimuovi l'elemento appena usato
    
    if('-us' in regex):
        all_separators = ['-us']
    
    if(r'\b(-----BEGIN RSA PRIVATE KEY----- [A-Za-z0-9+] -----END RSA PRIVATE KEY-----)\b' in regex):
        prefisso = "-----BEGIN RSA PRIVATE KEY----- "
        suffix_match = " -----END RSA PRIVATE KEY-----"

    if matches:
        # Usa una lista per associare la lunghezza successiva a quelli vuoti
        coppie = [
            (x, y if y else (matches[i + 1][1] if i + 1 < len(matches) else 'None'))
            for i, (x, y) in enumerate(matches)
            if x or (y or (i + 1 < len(matches) and matches[i + 1][1]))  # Rimuovi le coppie dove il primo valore è vuoto
        ]

        # Rimuovi le coppie con il primo valore vuoto ('') o con il secondo valore vuoto o solo spazi
        coppie = [pair for pair in coppie if pair[0] != '' and pair[1].strip() != '']

    else:
        #print("NON VA", regex)
        imperfect += 1
        aggiungi_regex_in_excel("regex.xlsx", regex)
        return None

    #charsets contiene tutti i possibili charset e le lunghezze di questi (opzionale)
    # Verifica se prefisso è una lista
    if isinstance(prefisso, list):  
        secret = ''.join(random.choice(prefisso))  # Sceglie un elemento casuale dalla lista
    else:
        secret = prefisso  # Usa direttamente prefisso se è una stringa

    charsets = []
    # Stampa le coppie risultanti
    if(not coppie):
        #print("NON VA", regex)
        imperfect += 1
        aggiungi_regex_in_excel("regex.xlsx", regex)
        return None
    else:
        for coppia in coppie:
            if(coppia[0] == '\\n\\r'):
                secret += ' '
            elif(coppia[0] == ' \\r\\n'):
                secret += ''
            else:
                charsets.append(coppia)
        for i, charset in enumerate(charsets):
            random_variable_part = generate_random_string(charset[0], charset[1])

            if(all_separators and len(all_separators)  == len(charsets) ):
                secret +=  all_separators[0]  # Usa l'elemento in testa
                all_separators.pop(0)  # Rimuovi l'elemento appena usato
            secret += random_variable_part
            # Aggiungi '-' solo se non è l'ultimo charset
            if i < len(charsets) - 1:
                if all_separators:
                    secret +=  all_separators[0]  # Usa l'elemento in testa
                    all_separators.pop(0)  # Rimuovi l'elemento appena usato
                else:
                    if (regex not in regex_no_separators):
                        secret += '-'
        
        if(suffix_match):
            for suffix in suffix_match:
                secret += suffix
        
        return secret

   
excel_file_path = "./Secret Regular Expression FILTERED.xlsx"
regex_df = pd.read_excel(excel_file_path, engine="openpyxl")

# Directory dove salvare i dati generati
output_dir = "./data/Secrets10"
os.makedirs(output_dir, exist_ok=True)

grouped_detectors = {}
for _, row in regex_df.iterrows():
    full_name = row['Secret Type']
    regex = row['Regular Expression']
    
    main_name = full_name.split(' ', 1)[0]
    if main_name not in grouped_detectors:
        grouped_detectors[main_name] = []
    grouped_detectors[main_name].append((full_name, regex))

for main_name, detectors in grouped_detectors.items():
    sanitized_main_name = sanitize_filename(main_name)
    all_secrets = []

    secret_types = [detector[0].split()[-1] for detector in detectors]
    generated_secrets = {secret_type: [] for secret_type in secret_types}

    for full_name, regex in detectors:
        secret_type = full_name.split()[-1]
        try:
            # Tentativo di generare 10 segreti per ogni tipo
            generated_secrets[secret_type] = [generate_secret(regex) for _ in range(10)]
        except Exception as e:
            # Se si verifica un errore, stampa un messaggio di errore e passa al prossimo detector
            print(f"Errore durante la generazione dei segreti per {full_name}: {e}")
            continue
    
    # Aggiungi i secret types e una riga di separazione
    all_secrets.append(" | ".join(secret_types))
    all_secrets.append("-" * len(" | ".join(secret_types)))

    # Aggiungi i segreti generati, evitando i valori None
    for i in range(10):
        row = [generated_secrets[secret_type][i] for secret_type in secret_types]
        # Rimuovi i valori None dalla riga
        row = [secret for secret in row if secret is not None]
        
        # Aggiungi la riga solo se non è vuota
        if row:
            all_secrets.append(" | ".join(row))

    # Scrivi i segreti nel file
    with open(os.path.join(output_dir, f"{sanitized_main_name}_secrets.txt"), "w") as file:
        file.write("\n".join(all_secrets))

print(f"Dataset generato con successo nella directory '{output_dir}'")
