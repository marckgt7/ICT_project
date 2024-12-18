import pandas as pd
import re
import os
import random
from openpyxl import Workbook, load_workbook

perfect = 0
imperfect = 0

# Liste per salvare i risultati
risultati_corte = []
risultati_lunghe = []
risultati_prefisso_mancante = []
risultati_charset_mancante = []
risultati_lunghezza_mancante = []

def aggiungi_regex_in_excel(file_excel, regex):
    try:
        # Controlla se il file Excel esiste
        try:
            workbook = load_workbook(file_excel)
            sheet = workbook.active
        except FileNotFoundError:
            # Crea un nuovo workbook se il file non esiste
            workbook = Workbook()
            sheet = workbook.active
            # Intestazione (opzionale)
            sheet.append(["Regex"])
        
        # Controlla se la regex è già presente
        existing_regexes = [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]
        if regex not in existing_regexes:
            # Aggiunge la regex
            sheet.append([regex])
            workbook.save(file_excel)
            print(f"Regex aggiunta: {regex}")
    except Exception as e:
        print(f"Errore: {e}")


def sanitize_filename(name):
    return re.sub(r'[\\/:*?"<>|]', '_', name)

def expand_charset(charset):
    expanded_charset = ""
    i = 0
    while i < len(charset):
        if i + 2 < len(charset) and charset[i + 1] == '-':
            start, end = charset[i], charset[i + 2]
            expanded_charset += ''.join(chr(c) for c in range(ord(start), ord(end) + 1))
            i += 3
        else:
            expanded_charset += charset[i]
            i += 1
    return expanded_charset


def generate_random_string(charset, length):
    global perfect, imperfect
    # Espandi il charset se contiene intervalli
    expanded_charset = expand_charset(charset)
    
    # Se length è un intervallo, scegli una lunghezza casuale nell'intervallo
    if ',' in length:
        min_len, max_len = length.split(',')
        min_len = int(min_len) if min_len else None
        max_len = int(max_len) if max_len else None
        
        if min_len is not None and max_len is not None:
            actual_length = random.randint(min_len, max_len)
        elif min_len is not None:
            actual_length = min_len  # Usa min_len come lunghezza fissa
        elif max_len is not None:
            actual_length = max_len  # Usa max_len come lunghezza fissa
        else:
            imperfect += 1

    elif length == 'None' or length == '\\':
        # Lunghezza casuale se è 'None'
        actual_length = random.randint(1, 100)  # Puoi cambiare l'intervallo
        imperfect += 1
    else:
        # Lunghezza fissa
        actual_length = int(length)
    
    #print(actual_length)
    return ''.join(random.choices(expanded_charset, k=actual_length))

# Funzione per analizzare una singola regex e ottenere informazioni dettagliate
def generate_secret(regex):
    global perfect, imperfect

    if regex == "EMAIL A PIACERE":
       return "testuser.1005@example.com"

    # Trova il prefisso
    match = re.search(r"\(\?:([a-zA-Z0-9|_.-]+)\)", regex) 
    if match:
        prefisso = match.group(1).split('|')
    else:
        match = re.search(r"\\b\(?([a-zA-Z0-9_\\-]+)\\?", regex) or \
                re.search(r'https://[^[]+', re.sub(r'\\', '', regex))
        if match:
            if match.lastindex is not None and match.lastindex >= 1: 
                prefisso = match.group(1) 
                prefisso = prefisso.replace("\\", "")
            else: 
                prefisso = match.group(0)  # HTTP 
        else:
            prefisso = ""
            aggiungi_regex_in_excel("regex.xlsx", regex)
            imperfect += 1

    # Trova il suffisso
    # Matching per suffissi tipo ".aha.io", ".jfrog.io" non chiusi tra [] o {}
    suffix_match = re.findall(r'(\.[a-zA-Z0-9\-]+|@[a-zA-Z0-9\-]+)', regex)
    if not suffix_match:
        suffix_match = re.findall(r'(?:\b|\W)([a-zA-Z0-9\-]+)(?=\)\\b)', regex)


    # Pattern per trovare i separatori (es. \. o \-)
    separator_pattern = r'\\([.=_|@\-])'  # Cattura solo il carattere dopo la barra rovesciata

    # Estrarre i separatori
    all_separators = []
    matches = re.findall(separator_pattern, regex)
    all_separators.extend(matches)

    # Se suffix_match ha qualcosa, rimuove gli ultimi n elementi da all_separators
    if suffix_match:
        n = len(suffix_match)
        all_separators = all_separators[:-n] if n <= len(all_separators) else []

    if '=' in all_separators:
        suffix_match.append('=')
        all_separators.remove('=')
    # Cerca charset e lunghezza
    #matches = re.findall(r"\[([a-zA-Z0-9\-_.=!@#$%^]+)\].*?\{(\d+),?(\d+)?\}", regex) or \
    #          re.findall(r"\[([^\]]+)\]\{(\d+),?(\d+)?\}", regex) 
    # Trova tutte le coppie tra parentesi quadre e graffe
    matches = re.findall(r"\[([^\]]+)\]|\{([^\}]+)\}", regex)

    if(".eyJ" in suffix_match):
        all_separators.insert(0, suffix_match[0])  # Usa l'elemento in testa
        suffix_match.pop(0)  # Rimuovi l'elemento appena usato

    if matches:

        # Usa una lista per associare la lunghezza successiva a quelli vuoti
        coppie = [
            (x, y if y else (matches[i + 1][1] if i + 1 < len(matches) else 'None'))
            for i, (x, y) in enumerate(matches)
            if x or (y or (i + 1 < len(matches) and matches[i + 1][1]))  # Rimuovi le coppie dove il primo valore è vuoto
        ]

        # Rimuovi le coppie con il primo valore vuoto ('') o con il secondo valore vuoto o solo spazi
        coppie = [pair for pair in coppie if pair[0] != '' and pair[1].strip() != '']

    else:
        #print("NON VA", regex)
        imperfect += 1
        aggiungi_regex_in_excel("regex.xlsx", regex)
        return None

    #charsets contiene tutti i possibili charset e le lunghezze di questi (opzionale)
    # Verifica se prefisso è una lista
    if isinstance(prefisso, list):  
        secret = ''.join(random.choice(prefisso))  # Sceglie un elemento casuale dalla lista
    else:
        secret = prefisso  # Usa direttamente prefisso se è una stringa

    charsets = []
    # Stampa le coppie risultanti
    if(not coppie):
        #print("NON VA", regex)
        imperfect += 1
        aggiungi_regex_in_excel("regex.xlsx", regex)
        return None
    else:
        for coppia in coppie:
            if(coppia[0] == '\\n\\r'):
                secret += ' '
            elif(coppia[0] == '\\r\\n'):
                secret += ' '
            else:
                charsets.append(coppia)

        for i, charset in enumerate(charsets):
            #print(f"Charset: {charset[0]}, Lunghezza: {charset[1]}")
            random_variable_part = generate_random_string(charset[0], charset[1])

            if(all_separators and len(all_separators)  == len(charsets) ):
                secret +=  all_separators[0]  # Usa l'elemento in testa
                all_separators.pop(0)  # Rimuovi l'elemento appena usato
            secret += random_variable_part
            # Aggiungi '-' solo se non è l'ultimo charset
            if i < len(charsets) - 1:
                if all_separators:
                    secret +=  all_separators[0]  # Usa l'elemento in testa
                    all_separators.pop(0)  # Rimuovi l'elemento appena usato
                else:
                    if (r'\b([a-z]{2}[0-9]{9})\b' not in regex):
                        secret += '-'
        
        
        
        if(suffix_match):
            for suffix in suffix_match:
                secret += suffix
        #print(regex)
        #print(charsets)
        #print(secret)
        
        return secret

   


# print("\nRegex analizate correttamente:", perfect)
# print("Problemi:", imperfect)

excel_file_path = "Secret Regular Expression MYNE.xlsx"
regex_df = pd.read_excel(excel_file_path, engine="openpyxl")

# Directory dove salvare i dati generati
output_dir = "generated_valid_secrets"
os.makedirs(output_dir, exist_ok=True)

grouped_detectors = {}
for _, row in regex_df.iterrows():
    full_name = row['Secret Type']
    regex = row['Regular Expression']
    
    main_name = full_name.split(' ', 1)[0]
    if main_name not in grouped_detectors:
        grouped_detectors[main_name] = []
    grouped_detectors[main_name].append((full_name, regex))

for main_name, detectors in grouped_detectors.items():
    sanitized_main_name = sanitize_filename(main_name)
    all_secrets = []

    secret_types = [detector[0].split()[-1] for detector in detectors]
    generated_secrets = {secret_type: [] for secret_type in secret_types}

    for full_name, regex in detectors:
        secret_type = full_name.split()[-1]
        try:
            # Tentativo di generare 10 segreti per ogni tipo
            generated_secrets[secret_type] = [generate_secret(regex) for _ in range(10)]
        except Exception as e:
            # Se si verifica un errore, stampa un messaggio di errore e passa al prossimo detector
            print(f"Errore durante la generazione dei segreti per {full_name}: {e}")
            continue
    
    # Aggiungi i secret types e una riga di separazione
    all_secrets.append(" | ".join(secret_types))
    all_secrets.append("-" * len(" | ".join(secret_types)))

    # Aggiungi i segreti generati, evitando i valori None
    for i in range(10):
        row = [generated_secrets[secret_type][i] for secret_type in secret_types]
        # Rimuovi i valori None dalla riga
        row = [secret for secret in row if secret is not None]
        
        # Aggiungi la riga solo se non è vuota
        if row:
            all_secrets.append(" | ".join(row))

    # Scrivi i segreti nel file
    with open(os.path.join(output_dir, f"{sanitized_main_name}_valid_secrets.txt"), "w") as file:
        file.write("\n".join(all_secrets))

print(f"Dataset generato con successo nella directory '{output_dir}'")
